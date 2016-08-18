#!/usr/bin/python
# coding: utf8

import os
import json
import datetime
import time
import smtplib
import subprocess
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import Encoders

from openpyxl import Workbook
from openpyxl.styles import colors, Font, Border, Side

DATABASE_FILE_NAME = '/home/lax/PycharmProjects/learn1/database4'
ADD_DATABASE_FILE_NAME = '/home/lax/PycharmProjects/learn1/add_database4'
EXCEL_TABLE_NAME = '/home/lax/PycharmProjects/learn1/new method test.xlsx'
SENDER_EADRESS = '***@gmail.com'
SENDER_EPASSWORD = '***'
RECEIVER_EADRESS = '***@gmail.com'
DATETIME_TEMPLATE = '%Y-%d-%m %H:%M:%S'


def get_sysresinfo(command):
    """ Executes console command and returns data as list """
    command_data = subprocess.Popen(command, stdout=subprocess.PIPE, shell=True)
    command_data, error_data = command_data.communicate()
    return command_data.replace(',', '.').split()


def get_cpu_info():
    cpu_console_data = get_sysresinfo('mpstat')
    cpu_user = float(cpu_console_data[21])
    cpu_sys = sum(float(x) for x in cpu_console_data[22:29])
    cpu_total = cpu_user + cpu_sys
    cpu_idle = float(cpu_console_data[30])
    return {'cpu_user': cpu_user, 'cpu_sys': cpu_sys, 'cpu_total': cpu_total, 'cpu_idle': cpu_idle}


def get_mem_info():
    mem_console_data = get_sysresinfo('free -m')
    mem_total, mem_used = int(mem_console_data[7]), int(mem_console_data[8])
    mem_free, mem_cached = int(mem_console_data[9]), int(mem_console_data[12])
    return {'mem_total': mem_total, 'mem_used': mem_used, 'mem_free': mem_free, 'mem_cached': mem_cached}


def get_hdd_info():
    hdd_console_data = get_sysresinfo('df -m --total')
    hdd_total, hdd_used = int(hdd_console_data[50]), int(hdd_console_data[51])
    hdd_free = int(hdd_console_data[52])
    return {'hdd_total': hdd_total, 'hdd_used': hdd_used, 'hdd_free': hdd_free}

###############################################################################################################


class SysinfoDatabase(object):
    """ Current database structure
    {timestamp (in seconds):  {
                               'cpu_user': 45.33,
                               'cpu_sys': 17.22,
                               'cpu_total': 60.11,
                               ...
                               'hdd_free':5952
                               }
    } """
    def __init__(self, db_file_name, datetime_format):
        self.db_file_name = db_file_name
        self.datetime_format = datetime_format
        if not os.path.isfile(self.db_file_name):
            self.sysinfo_database = {}
            with open(self.db_file_name, 'w') as self.database_file:
                self.database_file.write(json.dumps(self.sysinfo_database))
        else:
            with open(self.db_file_name, 'r') as self.database_file:
                self.sysinfo_database = self.database_file.read().strip()
                self.sysinfo_database = json.loads(self.sysinfo_database)

        self.db_index_timestamps = self.sysinfo_database.keys()
        self.db_index_timestamps.sort(reverse=True)
        self.sysinfo_keywords = ['cpu_user', 'cpu_sys', 'cpu_total', 'cpu_idle', 'mem_total', 'mem_used',
                                 'mem_free', 'mem_cached', 'hdd_total', 'hdd_used', 'hdd_free']

    def get_last_record_hour(self):  # Returns hour of last record in database
        return datetime.datetime.fromtimestamp(float(self.db_index_timestamps[0])).hour

    def select(self, start=None, end=None, hour_periods_limit=12):  # Select from database
        select_result = {}
        if not self.sysinfo_database:
            return select_result
        if start is None:
            start = self.db_index_timestamps[0]
        else:
            start = str(time.mktime(datetime.datetime.strptime(start, self.datetime_format).timetuple()))
        if end is None:
            end = self.db_index_timestamps[len(self.db_index_timestamps) - 1]
        else:
            end = str(time.mktime(datetime.datetime.strptime(end, self.datetime_format).timetuple()))
        if hour_periods_limit < 0:
            hour_periods_limit = 12

        current_period_timestamp = []
        periods_in_select_result = 0

        for index_timestamp in self.db_index_timestamps:
            if start >= index_timestamp >= end:
                index_timestamp_asobject = datetime.datetime.fromtimestamp(float(index_timestamp))
                if current_period_timestamp != index_timestamp_asobject.strftime('%Y,%m,%d,%H'):
                    current_period_timestamp = index_timestamp_asobject.strftime('%Y,%m,%d,%H')
                    periods_in_select_result += 1
                select_result[index_timestamp_asobject] = self.sysinfo_database[index_timestamp]
            if periods_in_select_result >= hour_periods_limit:
                break
            else:
                if index_timestamp < end:
                    break
        return select_result

    def average(self, input_data, groupbyhour=True):
        input_data_timestamps = input_data.keys()
        input_data_timestamps.sort(reverse=True)

        avg_period_result = {'cpu_user': 0, 'cpu_sys': 0, 'cpu_total': 0, 'cpu_idle': 0,
                             'mem_total': 0, 'mem_used': 0, 'mem_free': 0, 'mem_cached': 0,
                             'hdd_total': 0, 'hdd_used': 0, 'hdd_free': 0
                             }
        average_result = {}
        periods_in_result = 0
        averaged_in_period = 0
        period_timestamp = None

        for in_data_timestamp in input_data_timestamps:
            if not period_timestamp:
                period_timestamp = in_data_timestamp.strftime('%Y,%m,%d,%H')

            if groupbyhour:
                if period_timestamp != in_data_timestamp.strftime('%Y,%m,%d,%H'):
                    for key in self.sysinfo_keywords:
                        avg_period_result[key] /= averaged_in_period
                    average_result[datetime.datetime.strptime(period_timestamp, '%Y,%m,%d,%H')] = avg_period_result
                    periods_in_result += 1
                    averaged_in_period = 0
                    avg_period_result = {'cpu_user': 0, 'cpu_sys': 0, 'cpu_total': 0, 'cpu_idle': 0,
                                         'mem_total': 0, 'mem_used': 0, 'mem_free': 0, 'mem_cached': 0,
                                         'hdd_total': 0, 'hdd_used': 0, 'hdd_free': 0
                                         }
                    period_timestamp = in_data_timestamp.strftime('%Y,%m,%d,%H')
            single_db_record = input_data[in_data_timestamp]
            for key in self.sysinfo_keywords:
                avg_period_result[key] += single_db_record[key]
            averaged_in_period += 1

        if averaged_in_period != 0:
            for key in self.sysinfo_keywords:
                avg_period_result[key] /= averaged_in_period
                average_result[datetime.datetime.strptime(period_timestamp, '%Y,%m,%d,%H')] = avg_period_result
            periods_in_result += 1
        return average_result, periods_in_result

    def new_record(self, timestamp, data):  # Adding new record into database
        self.sysinfo_database[time.mktime(timestamp.timetuple())] = data
        with open(self.db_file_name, 'w') as self.database_file:
            self.database_file.write(json.dumps(self.sysinfo_database))
        self.db_index_timestamps = self.sysinfo_database.keys()
        self.db_index_timestamps.sort(reverse=True)

    def erase(self):  # Database full erase
        self.sysinfo_database = {}
        with open(self.db_file_name, 'w') as self.database_file:
            self.database_file.write(json.dumps(self.sysinfo_database))

    def clean(self, size_limit=500):  # Cleans database from old records (default is 500 record limit)
        database_size = len(self.db_index_timestamps)
        while database_size > size_limit:
            del self.sysinfo_database[self.db_index_timestamps[database_size - 1]]
            database_size -= 1
        with open(self.db_file_name, 'w') as self.database_file:
            self.database_file.write(json.dumps(self.sysinfo_database))
        self.db_index_timestamps = self.sysinfo_database.keys()
        self.db_index_timestamps.sort(reverse=True)

###################################################################################################################


def start_html_table():
    return """
    <html>
        <head></head>
        <body>
            <h1>System hour usage stats</h1>
            <table border = '1'>
            """


def extend_html_table(e_html_table, data, timestamp):
    e_html_table += """
                <tr>
                    <td>Averaging period {11} {12}:00 - {12}:59</td>
                </tr>
                <tr>
                    <td>CPU</td>
                    <td>total:{0:.2f}</td>
                    <td>user:{1:.2f}</td>
                    <td>system:{2:.2f}</td>
                    <td>idle:{3:.2f}</td>
                </tr>
                <tr>
                    <td>Memory</td>
                    <td>total:{4:d}</td>
                    <td>used:{5:d}</td>
                    <td>free:{6:d}</td>
                    <td>cached:{7:d}</td>
                </tr>
                <tr>
                    <td>Hard disk drive</td>
                    <td>total:{8:d}</td>
                    <td>used:{9:d}</td>
                    <td>free:{10:d}</td>
                </tr>
    """.format(data['cpu_total'], data['cpu_user'], data['cpu_sys'],
               data['cpu_idle'], data['mem_total'], data['mem_used'],
               data['mem_free'], data['mem_cached'], data['hdd_total'],
               data['hdd_used'], data['hdd_free'], timestamp.date(),
               timestamp.hour)

    return e_html_table


def end_html_table(e_html_table):
    return e_html_table + """
                </tr>
            </table>
        </body>
    </html>
    """

######################################################################################################################


class ExcelTable(object):
    def __init__(self):
        self.new_workbook = Workbook()
        self.new_worksheet = self.new_workbook.active
        self.black_font = Font(color=colors.BLACK)
        self.blue_font = Font(color=colors.BLUE)
        self.red_font = Font(color=colors.RED)
        self.border_allthin = Border(left=Side(border_style='thin', color=colors.BLACK),
                                     right=Side(border_style='thin', color=colors.BLACK),
                                     bottom=Side(border_style='thin', color=colors.BLACK),
                                     top=Side(border_style='thin', color=colors.BLACK))
        self.border_LTthik = Border(left=Side(border_style='thick', color=colors.BLACK),
                                    bottom=Side(border_style='thin', color=colors.BLACK),
                                    top=Side(border_style='thick', color=colors.BLACK))
        self.border_Tthik = Border(bottom=Side(border_style='thin', color=colors.BLACK),
                                   top=Side(border_style='thick', color=colors.BLACK))
        self.border_TRthik = Border(right=Side(border_style='thick', color=colors.BLACK),
                                    bottom=Side(border_style='thin', color=colors.BLACK),
                                    top=Side(border_style='thick', color=colors.BLACK))
        self.border_Lthik = Border(left=Side(border_style='thick', color=colors.BLACK),
                                   right=Side(border_style='thin', color=colors.BLACK),
                                   bottom=Side(border_style='thin', color=colors.BLACK),
                                   top=Side(border_style='thin', color=colors.BLACK))
        self.border_Rthik = Border(left=Side(border_style='thin', color=colors.BLACK),
                                   right=Side(border_style='thick', color=colors.BLACK),
                                   bottom=Side(border_style='thin', color=colors.BLACK),
                                   top=Side(border_style='thin', color=colors.BLACK))
        self.border_LBthik = Border(left=Side(border_style='thick', color=colors.BLACK),
                                    right=Side(border_style='thin', color=colors.BLACK),
                                    bottom=Side(border_style='thick', color=colors.BLACK),
                                    top=Side(border_style='thin', color=colors.BLACK))
        self.border_Bthik = Border(left=Side(border_style='thin', color=colors.BLACK),
                                   right=Side(border_style='thin', color=colors.BLACK),
                                   bottom=Side(border_style='thick', color=colors.BLACK),
                                   top=Side(border_style='thin', color=colors.BLACK))
        self.border_BRthik = Border(left=Side(border_style='thin', color=colors.BLACK),
                                    right=Side(border_style='thick', color=colors.BLACK),
                                    bottom=Side(border_style='thick', color=colors.BLACK),
                                    top=Side(border_style='thin', color=colors.BLACK))

        self.active_cell = self.new_worksheet.cell(row=2, column=2)
        self.active_cell.value = 'System 12 hour usage stats'
        self.active_cell.font = Font(color=colors.GREEN, italic=True, size=16)
        self.table_row_index = 2

        self.toprow_border_style = [self.border_LTthik, self.border_Tthik, self.border_Tthik,
                                    self.border_Tthik, self.border_TRthik]
        self.middlerow_border_style = [self.border_Lthik, self.border_allthin, self.border_allthin,
                                       self.border_allthin, self.border_Rthik]
        self.botmrow_border_style = [self.border_LBthik, self.border_Bthik, self.border_Bthik,
                                     self.border_Bthik, self.border_BRthik]
        self.row1_data = []
        self.row2_data = []
        self.row3_data = []
        self.row4_data = []

    def table_data_update(self, data, timestamp):
        self.row1_data = ['Averaging period %s %s:00 - %s:59' % (timestamp.date(), timestamp.hour, timestamp.hour),
                          '', '', '', '']
        self.row2_data = ['CPU', 'total: %.2f' % (data['cpu_total']), 'user: %.2f' % (data['cpu_user']), 'system: %.2f'
                          % (data['cpu_sys']), 'idle: %.2f' % (data['cpu_idle'])]
        self.row3_data = ['Memory', 'total: %d' % (data['mem_total']), 'used: %d'
                          % (data['mem_used']), 'free: %d' % (data['mem_free']), 'cached: %d' % (data['mem_cached'])]
        self.row4_data = ['Hard disk drive', 'total: %d' % (data['hdd_total']), 'used: %d'
                          % (data['hdd_used']), 'free: %d' % (data['hdd_free']), '']

    def extend_helper(self, exthe_row_data, exthe_border_style, exthe_first_col_font):
        self.table_row_index += 1
        for x in range(0, 5):
            self.active_cell = self.new_worksheet.cell(row=self.table_row_index, column=x+2)
            self.active_cell.value = exthe_row_data[x]
            self.active_cell.border = exthe_border_style[x]
            if x == 0:
                self.active_cell.font = exthe_first_col_font
            else:
                self.active_cell.font = self.black_font

    def extend(self, data, timestamp):
        self.table_data_update(data, timestamp)  # Update table variables
        self.extend_helper(self.row1_data, self.toprow_border_style, self.blue_font)
        self.extend_helper(self.row2_data, self.middlerow_border_style, self.red_font)
        self.extend_helper(self.row3_data, self.middlerow_border_style, self.red_font)
        self.extend_helper(self.row4_data, self.botmrow_border_style, self.red_font)

    def save(self, filename):
        self.new_worksheet.row_dimensions[2].height = 20
        self.new_worksheet.column_dimensions['A'].width = 5
        self.new_worksheet.column_dimensions['B'].width = 18
        self.new_worksheet.column_dimensions['C'].width = 15
        self.new_worksheet.column_dimensions['D'].width = 15
        self.new_worksheet.column_dimensions['E'].width = 15
        self.new_worksheet.column_dimensions['F'].width = 15
        self.new_workbook.save(filename)

#######################################################################################################################


def send_email(attach_table, excel_file, sender_adress, sender_pass, receiver_adress):
    mail = MIMEMultipart()
    mail['Subject'] = 'Test message new select'
    mail['From'] = 'Python interpreter'
    mail['To'] = 'To Lax-T'

    em_table_part = MIMEText(attach_table, 'html')

    if excel_file is not None:
        em_excel_file = MIMEBase('application', 'octet-stream')
        with open(excel_file, 'rb') as ef:
            em_excel_file.set_payload(ef.read())
            Encoders.encode_base64(em_excel_file)
        em_excel_file.add_header('Content-Disposition', 'attachment', filename='stats.xlsx')
        mail.attach(em_excel_file)

    mail.attach(em_table_part)
    em_client = smtplib.SMTP_SSL('smtp.gmail.com', '465')
    em_client.ehlo()
    em_client.login(sender_adress, sender_pass)  # password deleted
    em_client.sendmail(sender_adress, receiver_adress, mail.as_string())  # e-mail deleted
    em_client.close()

######################################################################################################################


def load_additionad_database(adb_file_name):
    if os.path.isfile(adb_file_name):  # Check if additional database exists
        with open(adb_file_name, 'r') as a_database_file:
            a_database = a_database_file.read().strip()
            a_database = json.loads(a_database)
        return a_database

    else:
        with open(adb_file_name, 'w') as a_database_file:
            a_database = {'last_em_send_hour': 24,
                          'emails_sent': 0}
            a_database_file.write(json.dumps(a_database))
        return a_database


def update_additional_database(adb_file_name, a_database):
    with open(adb_file_name, 'w') as a_database_file:
        a_database_file.write(json.dumps(a_database))

######################################################################################################################

if __name__ == '__main__':
    system_usage_info = dict(get_cpu_info().items() + get_mem_info().items() + get_hdd_info().items())
    system_time = datetime.datetime.now()

    systeminfo_database = SysinfoDatabase(DATABASE_FILE_NAME, DATETIME_TEMPLATE)
    select_result = systeminfo_database.select()
    averaging_result, periods_in_avg_result = systeminfo_database.average(select_result)

    additional_database = load_additionad_database(ADD_DATABASE_FILE_NAME)
    last_em_send_hour = additional_database['last_em_send_hour']
    emails_sent = additional_database['emails_sent']
    current_system_hour = system_time.hour

    html_table = start_html_table()

    if last_em_send_hour != current_system_hour+1 and periods_in_avg_result >= 1:  # check if averaging period changed and need to send email
        avg_res_timestamps = averaging_result.keys()
        avg_res_timestamps.sort(reverse=True)
        for index, timestamp_key in enumerate(avg_res_timestamps):  # 5 - table size limit
            html_table = extend_html_table(html_table, averaging_result[timestamp_key], timestamp_key)
            if index >= 4:
                break
        html_table = end_html_table(html_table)
        if emails_sent >= 11:  # 11 - is to include excel table in every 12th email (every 12 hours)
            new_excel_table = ExcelTable()
            for index, timestamp_key in enumerate(avg_res_timestamps):  # 12 - table size limit
                new_excel_table.extend(averaging_result[timestamp_key], timestamp_key)
                if index >= 11:
                    break
            new_excel_table.save(EXCEL_TABLE_NAME)
            send_email(html_table, EXCEL_TABLE_NAME, SENDER_EADRESS, SENDER_EPASSWORD, RECEIVER_EADRESS)
            emails_sent = 0
        else:
            send_email(html_table, None, SENDER_EADRESS, SENDER_EPASSWORD, RECEIVER_EADRESS)
            emails_sent += 1

        additional_database['emails_sent'] = emails_sent
        additional_database['last_em_send_hour'] = current_system_hour
        update_additional_database(ADD_DATABASE_FILE_NAME, additional_database)

    systeminfo_database.new_record(system_time, system_usage_info)
    systeminfo_database.clean()
